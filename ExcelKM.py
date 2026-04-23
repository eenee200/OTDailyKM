import requests
import json
import pandas as pd
import numpy as np
from math import radians, sin, cos, sqrt, atan2
from datetime import datetime, timedelta
import os
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.mime.text import MIMEText
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import openpyxl.utils


def haversine_distance(lat1, lon1, lat2, lon2):
    """
    Calculate the great circle distance between two points 
    on the earth (specified in decimal degrees)
    Returns distance in kilometers
    """
    lat1, lon1, lat2, lon2 = map(radians, [lat1, lon1, lat2, lon2])
    
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
    c = 2 * atan2(sqrt(a), sqrt(1-a))
    
    r = 6371  # Radius of earth in kilometers
    
    return r * c


def fetch_gps_data_from_api(server_domain, api_key, object_id, start_date, end_date):
    """
    Fetch GPS data from API for a single vehicle
    """
    api_url = f"{server_domain}/api/api.php?api=user&key={api_key}&cmd=OBJECT_GET_MESSAGES,{object_id},{start_date},{end_date}"
    
    try:
        print(f"Fetching GPS data for object {object_id}...")
        response = requests.get(api_url)
        response.raise_for_status()
        
        data = json.loads(response.text)
        
        if isinstance(data, list) and len(data) > 0:
            coordinates = []
            for entry in data:
                if isinstance(entry, list) and len(entry) >= 3:
                    try:
                        point = {
                            'timestamp': datetime.strptime(entry[0], '%Y-%m-%d %H:%M:%S')+ timedelta(hours=8),
                            'latitude': float(entry[1]),
                            'longitude': float(entry[2])
                        }
                        coordinates.append(point)
                    except (ValueError, IndexError) as e:
                        continue
            
            df = pd.DataFrame(coordinates)
            print(f"  Fetched {len(df)} GPS points")
            return df
        else:
            print(f"  No data returned")
            return pd.DataFrame()
            
    except Exception as e:
        print(f"  Error fetching data: {e}")
        return pd.DataFrame()


def calculate_gps_distances(df):
    """
    Calculate distances between consecutive GPS points using Haversine formula
    Returns DataFrame with distance and time_diff columns added
    """
    if len(df) < 2:
        df['distance_km'] = 0
        df['time_diff_hours'] = 0
        return df
    
    df = df.sort_values('timestamp').reset_index(drop=True)
    
    distances = []
    time_diffs = []
    
    for i in range(len(df)):
        if i == 0:
            distances.append(0)
            time_diffs.append(0)
        else:
            # Calculate distance
            dist = haversine_distance(
                df.loc[i-1, 'latitude'], df.loc[i-1, 'longitude'],
                df.loc[i, 'latitude'], df.loc[i, 'longitude']
            )
            
            # Calculate time difference in hours
            time_diff = (df.loc[i, 'timestamp'] - df.loc[i-1, 'timestamp']).total_seconds() / 3600
            
            distances.append(dist)
            time_diffs.append(time_diff)
    
    df['distance_km'] = distances
    df['time_diff_hours'] = time_diffs
    
    return df


def filter_unrealistic_movements(df, max_time_gap=1.0, max_speed_kmh=200.0, min_movement_km=0.01):
    """
    Filter out unrealistic GPS movements
    - max_time_gap: Maximum time between points in hours (default 1 hour)
    - max_speed_kmh: Maximum realistic speed in km/h (default 200 km/h)
    - min_movement_km: Minimum movement to consider (filters GPS drift)
    """
    if len(df) < 2:
        return df
    
    initial_count = len(df)
    
    # Calculate implied speed
    df['implied_speed'] = df.apply(
        lambda row: (row['distance_km'] / row['time_diff_hours']) if row['time_diff_hours'] > 0 else 0,
        axis=1
    )
    
    # Filter conditions
    valid_mask = (
        (df['time_diff_hours'] <= max_time_gap) &  # Not too much time gap
        (df['implied_speed'] <= max_speed_kmh) &    # Realistic speed
        ((df['distance_km'] >= min_movement_km) | (df.index == 0))  # Real movement (or first point)
    )
    
    df_filtered = df[valid_mask].copy()
    
    filtered_count = len(df_filtered)
    print(f"  Filtered movements: {initial_count} -> {filtered_count} ({initial_count - filtered_count} removed)")
    
    return df_filtered


def calculate_hourly_distance(df, vehicle_name):
    """
    Calculate hourly distances from GPS data
    """
    if len(df) == 0:
        return pd.DataFrame()
    
    # Add date and hour columns
    df['date'] = df['timestamp'].dt.date
    df['hour'] = df['timestamp'].dt.hour
    
    # Group by date and hour, sum distances
    hourly = df.groupby([df['timestamp'].dt.floor('h')])['distance_km'].sum().reset_index()
    hourly = hourly.rename(columns={'timestamp': 'timestamp', 'distance_km': 'distance_km'})
    
    # Add vehicle name
    hourly['vehicle'] = vehicle_name
    
    # Filter out negligible movements
    hourly = hourly[hourly['distance_km'] > 0.01]
    
    # Round distances
    hourly['distance_km'] = hourly['distance_km'].round(2)
    
    return hourly[['vehicle', 'timestamp', 'distance_km']]


def process_multiple_vehicles(vehicles_config):
    """
    Process multiple vehicles and combine their hourly distances
    
    vehicles_config: list of dicts with keys:
        - name: vehicle name
        - object_id: GPS tracker object ID
        - server_domain: API server domain
        - api_key: API key
        - start_date: start date (YYYY-MM-DD HH:MM:SS)
        - end_date: end date (YYYY-MM-DD HH:MM:SS)
    """
    all_hourly_distances = []
    
    for vehicle in vehicles_config:
        print(f"\nProcessing {vehicle['name']}...")
        
        # Fetch GPS data
        df = fetch_gps_data_from_api(
            vehicle['server_domain'],
            vehicle['api_key'],
            vehicle['object_id'],
            vehicle['start_date'],
            vehicle['end_date']
        )
        
        if len(df) == 0:
            print(f"  No data for {vehicle['name']}, skipping")
            continue
        
        # Calculate distances
        df = calculate_gps_distances(df)
        
        
        # Calculate hourly distances
        hourly = calculate_hourly_distance(df, vehicle['name'])
        
        if len(hourly) > 0:
            total_distance = hourly['distance_km'].sum()
            print(f"  Total distance: {total_distance:.2f} km")
            all_hourly_distances.append(hourly)
    
    if not all_hourly_distances:
        raise ValueError("No valid vehicle data found")
    
    combined_df = pd.concat(all_hourly_distances, ignore_index=True)
    return combined_df


def save_to_excel_time_format(hourly_distance, output_file):
    """
    Save hourly distance data for multiple vehicles to an Excel file
    """
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    # Format the hour column to range format (e.g., "00:00-01:00", ..., "23:00-00:00")
    hourly_distance['Date'] = hourly_distance['timestamp'].dt.date
    hourly_distance['Hour'] = hourly_distance['timestamp'].dt.hour.apply(
        lambda h: f"{str(h).zfill(2)}:00-{str((h+1)%24).zfill(2)}:00"
    )
    
    # Prepare all unique dates and vehicles
    all_vehicles = sorted(hourly_distance['vehicle'].unique())
    
    # Generate complete date range
    existing_dates = sorted(hourly_distance['Date'].unique())
    
    all_dates = []
    if existing_dates:
        start_date = existing_dates[0]
        end_date = existing_dates[-1]
        
        current_date = start_date
        while current_date <= end_date:
            all_dates.append(current_date)
            current_date += timedelta(days=1)
    
    # Prepare hours (4:00 to 20:00 as in your original code)
    all_hours = [f"{str(h).zfill(2)}:00-{str((h+1)%24).zfill(2)}:00" for h in range(4, 21)]
    
    # Define styles
    grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    font_style = Font(name="Aptos Narrow", size=11)
    align_center = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        sheet_names = []
        
        # Process each date into its own sheet
        for date in all_dates:
            daily_data = hourly_distance[hourly_distance['Date'] == date]
            
            # Create DataFrame with all hours and vehicles
            all_combinations = pd.DataFrame(
                [(vehicle, hour) for vehicle in all_vehicles for hour in all_hours],
                columns=['vehicle', 'Hour']
            )
            all_combinations['distance_km'] = 0
            
            # Merge with actual data
            if not daily_data.empty:
                daily_data = daily_data.merge(all_combinations, on=['vehicle', 'Hour'], how='right')
                daily_data['distance_km'] = daily_data['distance_km_x'].fillna(daily_data['distance_km_y'])
                daily_data = daily_data[['vehicle', 'Hour', 'distance_km']]
            else:
                daily_data = all_combinations
            
            # Pivot
            daily_pivot = daily_data.pivot(index='vehicle', columns='Hour', values='distance_km').fillna(0)
            daily_pivot = daily_pivot.reindex(columns=all_hours, fill_value=0)
            daily_pivot = daily_pivot.reindex(all_vehicles, fill_value=0)
            
            sheet_name = str(date)
            daily_pivot.to_excel(writer, sheet_name=sheet_name)
            sheet_names.append(sheet_name)
        
        # Calculate total distance
        total_distance = hourly_distance.groupby('vehicle')['distance_km'].sum().reset_index()
        total_distance = total_distance.rename(columns={'distance_km': 'Total Distance (km)'})
        
        full_total_distance = pd.DataFrame({'vehicle': all_vehicles})
        full_total_distance = full_total_distance.merge(total_distance, on='vehicle', how='left')
        full_total_distance['Total Distance (km)'] = full_total_distance['Total Distance (km)'].fillna(0)
        
        total_sheet_name = "Total"
        full_total_distance.to_excel(writer, sheet_name=total_sheet_name, index=False)
        sheet_names.append(total_sheet_name)
        
        # Apply formatting
        workbook = writer.book
        for sheet_name in sheet_names:
            worksheet = workbook[sheet_name]
            
            worksheet.freeze_panes = worksheet["B2"]
            
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.font = font_style
                    cell.alignment = align_center
                    cell.border = thin_border
            
            for cell in worksheet["A"]:
                cell.fill = grey_fill
            
            for col in worksheet.columns:
                max_length = max(len(str(cell.value)) for cell in col)
                worksheet.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = max_length + 2
        
        # Reorder sheets
        def sheet_sort_key(ws):
            try:
                return datetime.strptime(ws.title, "%Y-%m-%d")
            except ValueError:
                return datetime.max
        
        workbook._sheets.sort(key=sheet_sort_key)


def send_email_with_attachment(sender_email, sender_password, receiver_emails, 
                             subject, message, attachment_path):
    """
    Send email with Excel report attachment to multiple recipients
    
    Parameters:
    -----------
    sender_email : str
        Email address of the sender
    sender_password : str
        Password for the sender's email account
    receiver_emails : list or str
        List of recipient email addresses or a single email address
    subject : str
        Subject of the email
    message : str
        Body of the email
    attachment_path : str
        Path to the file to be attached
    """
    # Convert single email to list if necessary
    if isinstance(receiver_emails, str):
        receiver_emails = [receiver_emails]
    
    # Remove any empty strings and whitespace
    receiver_emails = [email.strip() for email in receiver_emails if email.strip()]
    
    if not receiver_emails:
        print("No valid receiver emails provided")
        return False
    
    try:
        # Create the email message
        email_message = MIMEMultipart()
        email_message['From'] = sender_email
        email_message['To'] = ', '.join(receiver_emails)  # Join all recipients with commas
        email_message['Subject'] = subject
        
        # Attach the message body
        email_message.attach(MIMEText(message, 'plain', 'utf-8'))
        
        # Attach the file
        with open(attachment_path, 'rb') as file:
            part = MIMEApplication(file.read(), Name=os.path.basename(attachment_path))
            part['Content-Disposition'] = f'attachment; filename="{os.path.basename(attachment_path)}"'
            email_message.attach(part)
        
        # Send the email
        with smtplib.SMTP("smtp.gmail.com", 587) as server:
            server.starttls()
            server.login(sender_email, sender_password)
            server.send_message(email_message)
        
        print(f"Email sent successfully to: {', '.join(receiver_emails)}")
        return True
    
    except Exception as e:
        print(f"Error sending email: {e}")
        return False



if __name__ == "__main__":
    # ==========================================================================
    # CONFIGURATION - Set these once
    # ==========================================================================

    first = datetime.now().replace( hour=16, minute=0, second=0, microsecond=0) - timedelta (days=1)
    prev = first - timedelta (days=1)
    SERVER_DOMAIN = "https://fms.gpsbox.mn"  # Your server domain
    API_KEY = "22E036123BFB94ADC793699001B25D6C"  # Your API key
    START_DATE = prev  # Start date
    END_DATE = first # End date

    # for the description and some attechment sake

    Dprev = prev + timedelta(hours=8)
    Dfirst = first + timedelta(hours=8)
    
    # ==========================================================================
    # VEHICLES - Only add object IDs and names here
    # ==========================================================================
    vehicles = [
        ('BU-5373', '358480089964961'),
        ('BU-5509', '863719064005110'),
      
        ('BU-5360', '359633104788653'),
        ('BU-5426', '860896052071727'),
        ('BU-5428', '359633104790329'),
        ('BU-5429', '860896052112265'),
        ('BU-5435', '359632109418183'),
        ('BU-5441', '860896052102225'),

        ('BU-5450', '350612076904307'),

        ('BU-5452', '860896052104726'),
        ('BU-5459', '860896052070877'),
        ('BU-5464', '359633100482681'),
        ('BU-5469', '860896051934115'),
        ('BU-5471', '860896052102522'),
        ('BU-5475', '860896052113099'),
        ('BU-5493', '359632109548658'),

        ('BU-5494', '860896050898923'),

        ('BU-5495', '359633107543220'),

        ('BU-5497', '860896051265700'),
        ('BU-5498', '860896050898956'),
        ('BU-5501', '350612076904299'),
        ('BU-5505', '354018112879253'),

        ('BU-5511', '860896050935733'),
        ('BU-5512', '860896052101417'),
        ('BU-5513', '860896052108255'),
        ('BU-5527', '860896050930809'),
        ('BU-5528', '352592572341600'),
        ('BU-5529', '860896051252674'),

        ('BU-5535', '866381054879504'),
        ('BU-5537', '359633104788521'),
        ('BU-5542', '354018112819176'),
        ('BU-5543', '354018112819630'),
        ('BU-5544', '863719064005052'),
        ('BU-5546', '860896052078888'),
        ('BU-5547', '359633107541869'),

        ('BU-5548', '359633107542412'),
        ('BU-5549', '359633107531837'),
        ('BU-5550', '359633107534054'),
        ('BU-5551', '352592572401115'),
        ('BU-5553', '860896052080009'),
        ('BU-5555', '860896052078854'),
        ('BU-5556', '359633107533817'),
        ('BU-5558', '350612076248523'),

        ('BU-5559', '352016708460258'),
        ('BU-5560', '352592572401099'),
        ('BU-5561', '863719069379031'),
        ('BU-5562', '863719062240743'),
        ('BU-5564', '863719069343441'),
        ('BU-5565', '863719069343557'),
        ('BU-5566', '863719069379023'),
        ('BU-5567', '863719069378611'),

        ('BU-5569', '863719069344514'),
        ('BU-5570', '860896051268746'),
        ('BU-5571', '359633107542396'),

        ('7913 ӨМЕ (DZ)', '352592572389781'),
        ('BU-5457', '352592572347268'),
        ('BU-5474', '359633107532223'),
        ('BU-5508', '860896052099629'),

        ('BU-5502', '860896051239945'),
        ('BU-5503', '860896051267029'),
        ('BU-5530', '352592572386027'),

        ('BU-5504', '358480089964961'),

        ('BU-8011' , '865124070237161'),
        ('BU-8012' , '358480087922417'),
        ('BU-8014' , '866381054964066'),
        ('BU-8015' , '860896050899137'),
        ('BU-8024' , '865124070237146'),
        ('BU-8028' , '860896050898972'),
        ('BU-8044' , '358480088010717'),
        ('BU-8045' , '358480087927416'),
        ('BU-8062' , '863719064005029'),
        ('BU-8063' , '863719069272236'),
        ('BU-9061' , '354018114841749'),
        ('BU-9062' , '863719069272939'),
        ('BU-9063' , '354018112717156'),
        ('BU-9064' , '352592572386282'),
        ('BU-9065' , '866381054964140'),
        ('BU-9066' , '865124070239274'),
        ('BU-9067' , '359633104788539'),
        ('BU-9068' , '359632108679983'),
        ('BU-9069' , '865124070239100'),
        ('BU-9070' , '354018112831510'),
        ('BU-9072' , '359633104789263'),
        ('BU-9073' , '352592572347318'),
        ('BU-9074' , '359633104711622'),
        ('BU-9075' , '352592572399582'),
        ('BU-9076' , '352592572386209'),
        ('BU-9077' , '358480087889509'),
        ('BU-9078' , '863719062208104'),
        ('LV-8022' , '865124070237070'),
        ('LV-8038' , '358480085118331'),
        ('LV-8039' , '358480085101279'),
        ('LV-8054' , '358480085134544'),
        ('LV-8056' , '358480087927390'),
        ('LV-8057' , '866069064378043'),
        ('LV-8058' , '866069064377995'),
        ('LV-8059' , '352016708288535'),
        ('LV-8061' , '866069069039905'),

    ]
    
    # Build configuration (no need to edit below)
    vehicles_config = [
        {
            'name': name,
            'object_id': obj_id,
            'server_domain': SERVER_DOMAIN,
            'api_key': API_KEY,
            'start_date': START_DATE,
            'end_date': END_DATE
        }
        for name, obj_id in vehicles
    ]
    
    try:
        # Process all vehicles
        hourly_distance = process_multiple_vehicles(vehicles_config)
        
        # Save to Excel
        output_file = "gps_hourly_mileage.xlsx"
        save_to_excel_time_format(hourly_distance, output_file)
        
        print("\n" + "="*70)
        print("SUMMARY")
        print("="*70)
        print(f"\nTotal vehicles processed: {len(hourly_distance['vehicle'].unique())}")
        print(f"Total distance for all vehicles: {hourly_distance['distance_km'].sum():.2f} km")
        print(f"\nResults saved to: {output_file}")

        SENDER_EMAIL = os.environ.get('SENDER_EMAIL')   
        SENDER_PASSWORD = os.environ.get('SENDER_PASSWORD')
        RECEIVER_EMAILS = "saranguam@riotinto.com, otjourneymanagement@ot.mn"

        if output_file:
                
                # Send email
                send_email_with_attachment(
                    SENDER_EMAIL,
                    SENDER_PASSWORD,
                    RECEIVER_EMAILS,
                    f"{Dprev} - {Dfirst} ны цаг бүрийн замын уртын тайлан",
                    f"ТТТоолс тайлан өдөр бүрийн",
                    output_file
                )
        else:
                print("No data was processed successfully for any vehicle.")
        
        
    except Exception as e:
        print(f"\nError: {str(e)}")
        import traceback
        traceback.print_exc()

    
